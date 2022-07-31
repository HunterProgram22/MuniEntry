"""Module for getting case data from Excel files."""
from types import MappingProxyType

from loguru import logger

from munientry.data.excel_functions import (
    create_headers_dict,
    get_excel_file_headers,
    load_active_worksheet,
)
from munientry.models.excel_models import CaseExcelData
from munientry.settings import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import Workbook


# Column Headers
COL_CASE = 'CaseNumber'
COL_DEF_LAST_NAME = 'DefLastName'
COL_DEF_FIRST_NAME = 'DefFirstName'
COL_CHARGE = 'ChargeDescription'
COL_STATUTE = 'SectionCode'
COL_DEGREE = 'DegreeCode'
COL_INSURANCE = 'InsuranceStatus'
COL_MOVING_OFFENSE = 'IsMoving'
COL_DEF_ATTY_LAST_NAME = 'AttorneyLastName'
COL_DEF_ATTY_FIRST_NAME = 'AttorneyFirstName'
COL_DEF_ATTY_TYPE = 'PubDef'

NO_DATA = 'No Data'
OFFENSE_CLEAN_DICT = MappingProxyType({
    'UCM': '',
    'M1': '',
    'M2': '',
    'M3': '',
    'M4': '',
    'MM': '',
    'PETTY': '',
    '(UCM)': '',
    'DUS': 'DUS',
    'OVI': 'OVI',
    'BMV': 'BMV',
    'OBMV': 'BMV',
    'FRA': 'FRA',
    'OL': 'OL',
    'OMVI': 'OVI',
    'FRA/JUDGMENT': 'FRA / Judgment',
    'OR': '/',
    'W/O': 'Without',
    'A': 'a',
    'TO': 'to',
    'SUSP': 'Suspension',
    '-': '',
    'OF': 'of',
    'IN': 'in',
    'AND': 'and',
})


def return_cases_data_from_excel(excel_file: str) -> list[CaseExcelData]:
    """Loads active worksheet, generates header dict, and creates case data list."""
    worksheet = load_active_worksheet(excel_file)
    header_list = get_excel_file_headers(worksheet)
    headers_dict = create_headers_dict(header_list)
    return create_case_data_list(worksheet, headers_dict)


def clean_statute_name(statute: str) -> str:
    """Removes trailing asteriks that are often part of data from AuthorityCourt."""
    return statute.rstrip('*')


def clean_offense_name(offense: str) -> str:
    """Sets offense name to title case (except for abbreviations) and removes degree of charge."""
    offense_word_list = offense.split()
    clean_offense_word_list = []
    for word in offense_word_list:
        if OFFENSE_CLEAN_DICT.get(word) is not None:
            clean_offense_word_list.append(OFFENSE_CLEAN_DICT.get(word))
            continue
        clean_offense_word_list.append(word.capitalize())
    return ' '.join([str(clean_word) for clean_word in clean_offense_word_list])


def create_case_data_list(worksheet: 'Workbook.active', headers_dict: dict) -> list[CaseExcelData]:
    """Returns a list of CaseExcelData objects.

    The CaseExcelData object is loaded with the data from each row (case) and the attributes
    of each case are mapped to the specific column that contains the case attribute. The load
    portion of the function starts with row 2 because row 1 is headers.
    """
    case_data_list: list = []
    for row in range(2, worksheet.max_row + 1):
        case = CaseExcelData()
        case.case_number = get_cell_value(worksheet, row, headers_dict[COL_CASE])
        case.defendant_last_name = (
            get_cell_value(worksheet, row, headers_dict[COL_DEF_LAST_NAME]).title()
        )
        case.defendant_first_name = (
            get_cell_value(worksheet, row, headers_dict[COL_DEF_FIRST_NAME]).title()
        )

        offense = get_cell_value(worksheet, row, headers_dict[COL_CHARGE])
        case.offense = clean_offense_name(offense)

        statute = get_cell_value(worksheet, row, headers_dict[COL_STATUTE])
        case.statute = clean_statute_name(statute)

        case.degree = get_cell_value(worksheet, row, headers_dict[COL_DEGREE])
        case.fra_in_file = get_cell_value(worksheet, row, headers_dict[COL_INSURANCE])
        case.moving_bool = get_cell_value(worksheet, row, headers_dict[COL_MOVING_OFFENSE])
        case.def_atty_last_name = (
            get_cell_value(worksheet, row, headers_dict[COL_DEF_ATTY_LAST_NAME]).title()
        )
        case.def_atty_first_name = (
            get_cell_value(worksheet, row, headers_dict[COL_DEF_ATTY_FIRST_NAME]).title()
        )
        case.def_atty_type = get_cell_value(worksheet, row, headers_dict[COL_DEF_ATTY_TYPE])
        case_data_list.append(case)
    return case_data_list


def get_cell_value(worksheet: 'Workbook.active', row: int, col: int) -> str:
    """Returns the cell value for a cell in the active excel worksheet."""
    if worksheet.cell(row=row, column=col).value is None:
        return set_cell_value_if_none(worksheet, col)
    return worksheet.cell(row=row, column=col).value


def set_cell_value_if_none(worksheet: 'Workbook.active', col: int) -> str:
    """Returns a specific string for certain attributes if there is no data."""
    if worksheet.cell(row=1, column=col).value == COL_INSURANCE:
        return 'U'
    if worksheet.cell(row=1, column=col).value == COL_DEF_ATTY_LAST_NAME:
        return ''
    if worksheet.cell(row=1, column=col).value == COL_DEF_ATTY_FIRST_NAME:
        return ''
    return NO_DATA


if __name__ == '__main__':
    logger.log('IMPORT', f'{__name__} run directly.')
else:
    logger.log('IMPORT', f'{__name__} imported.')
