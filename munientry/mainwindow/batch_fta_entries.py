"""Module for creating a batch of Failure to Appear entries."""
from datetime import datetime
from dataclasses import asdict, dataclass

from docxtpl import DocxTemplate
from loguru import logger
from openpyxl import Workbook, load_workbook

from munientry.data.loader_functions import (
    create_headers_dict,
    get_excel_file_headers,
    load_active_worksheet,
)
from munientry.settings import DB_PATH, SAVE_PATH, TEMPLATE_PATH

NO_DATA = 'No Data'
COL_CASE_NUMBER = 'CaseNumber'
COL_CASE_TYPE = 'CaseTypeCode'
COL_EVENT_DATE = 'CaseEventDate'
COL_DEF_LAST_NAME = 'DefLastName'
COL_DEF_FIRST_NAME = 'DefFirstName'


def create_entry(case_data):
    """General create entry function that populates a template with data."""
    doc = DocxTemplate(fr'{TEMPLATE_PATH}\Batch_Failure_To_Appear_Arraignment_Template.docx')
    doc.render(case_data.get_case_information())
    docname = set_document_name(case_data.case_number)
    doc.save(fr'{SAVE_PATH}\batch\{docname}')


def set_document_name(case_number: str) -> str:
    """Sets document name that includes case number."""
    return f'{case_number}_FTA_Arraignment.docx'


@dataclass
class BatchCaseInformation:
    case_number: str  = None
    case_type_code: str = None
    warrant_rule: str = None
    case_event_date: str = None
    def_first_name: str = None
    def_last_name: str = None

    def get_case_information(self):
        return asdict(self)


def get_case_list_from_excel(excel_file: str) -> list[BatchCaseInformation]:
    """Loads active worksheet, generates header dict, and list with case data objects."""
    worksheet = load_active_worksheet(excel_file)
    header_list = get_excel_file_headers(worksheet)
    headers_dict = create_headers_dict(header_list)
    return create_batch_case_list(worksheet, headers_dict)


def create_batch_case_list(worksheet: Workbook.active, header_dict: dict) -> list[BatchCaseInformation]:
    """Reads through an excel file and gets case data."""
    batch_case_data: list = []
    row_count = worksheet.max_row + 1
    for row in range(2, row_count):
        case_number = get_cell_value(worksheet, row, header_dict[COL_CASE_NUMBER])
        case_type_code = get_cell_value(worksheet, row, header_dict[COL_CASE_TYPE])
        warrant_rule = set_warrant_rule(case_type_code)
        case_event_date = get_cell_value(worksheet, row, header_dict[COL_EVENT_DATE])
        case_event_date = format_event_date(case_event_date)
        def_first_name = get_cell_value(worksheet, row, header_dict[COL_DEF_FIRST_NAME])
        def_first_name = def_first_name.title()
        def_last_name = get_cell_value(worksheet, row, header_dict[COL_DEF_LAST_NAME])
        def_last_name = def_last_name.title()
        case_information = BatchCaseInformation(
            case_number,
            case_type_code,
            warrant_rule,
            case_event_date,
            def_first_name,
            def_last_name,
        )
        batch_case_data.append(case_information)
    return batch_case_data


def get_cell_value(ws: Workbook.active, row: int, col: int) -> str:
    """Returns the cell value for a cell in the active excel worksheet."""
    if ws.cell(row=row, column=col).value is None:
        return set_cell_value_if_none(ws, col)
    return ws.cell(row=row, column=col).value


def set_cell_value_if_none(ws: Workbook.active, col: int) -> str:
    """Returns a specific string for certain attributes if there is no data."""
    return NO_DATA


def set_warrant_rule(case_type_code: str) -> str:
    """Returns a string with the specific warrant rule based on the case type."""
    if case_type_code == 'CRB':
        return 'Criminal Rule 4'
    return 'Traffic Rule 7'


def format_event_date(case_event_date: datetime) -> str:
    return case_event_date.strftime('%B %d, %Y')


def run_batch_fta_arraignments() -> int:
    batch_case_list = get_case_list_from_excel(f'{DB_PATH}\\Batch_FTA_Arraignments.xlsx')
    entry_count = 0
    for index, case in enumerate(batch_case_list):
        logger.info(f'Creating Batch FTA entry for: {batch_case_list[index].case_number}')
        create_entry(batch_case_list[index])
        entry_count +=1
    return entry_count


if __name__ == '__main__':
    run_batch_fta_arraignments()
    logger.info(f'{__name__} run directly.')
else:
    logger.info(f'{__name__} imported.')
