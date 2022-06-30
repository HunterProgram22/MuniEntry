"""Module containing all classes and functions related to creation and access to
the databases used by the application."""

from abc import ABC, abstractmethod
import os
import sys

from loguru import logger
from openpyxl import load_workbook  # type: ignore
from PyQt5.QtSql import QSqlQuery, QSqlDatabase
from package.excel_loaders.case_excel_loader import return_cases_data_from_excel
from package.excel_loaders.charge_excel_loader import create_charges_data_list

from settings import DB_PATH, CHARGES_TABLE, EXCEL_DAILY_CASE_LISTS
from package.models.cms_models import CmsCaseInformation
from package.database_controllers.sql_queries import (
    create_daily_case_list_tables_sql_query,
    create_charges_table_sql_query,
    insert_daily_case_list_tables_sql_query,
    insert_charges_sql_query,
    delete_table_sql_query,
    select_case_data_sql_query,
    select_distinct_offense_statute_sql_query,
    select_distinct_def_last_def_first_case_number_sql_query,
    select_statute_from_charges_for_offense_type_sql_query,
)


class CaseSQLRetriever(ABC):
    @abstractmethod
    def query_case_data(self):
        raise NotImplementedError

    @abstractmethod
    def load_data_into_case(self):
        raise NotImplementedError

    @abstractmethod
    def load_case(self):
        raise NotImplementedError


class CriminalCaseSQLRetriever(CaseSQLRetriever):
    """Class for getting case data from a SQL database and packaging it into an object to be
    loaded into the application.

    :case_number: The selected case number from the case selected in the daily case lists box
    on the main window of the application.

    :case_table: The string name of the daily case list box that the case is selected from. Used
    to select the table in the database from which to obtain the case."""

    def __init__(self, case_number: str, case_table: str) -> None:
        self.case_number = case_number
        self.case_table = case_table
        self.abbreviation_list = ["DUS", "OVI", "BMV", "FRA", "OL"]
        self.delete_word_list = ["UCM", "M1", "M2", "M3", "M4", "MM", "PETTY", "(UCM)"]
        self.database = open_db_connection("con_daily_case_lists")
        self.case = CmsCaseInformation()
        self.query_case_data()
        self.load_data_into_case()
        self.query.finish()
        close_db_connection(self.database)

    def query_case_data(self) -> None:
        """Query database based on cms_case number to return the data to load for the dialog.
        Current - Query.value(0) is id, then 1 is case_number, 2 is last_name, 3 is first_name."""
        query_string = select_case_data_sql_query(self.case_table, self.case_number)
        self.query = QSqlQuery(self.database)
        self.query.prepare(query_string)
        self.query.bindValue(self.case_number, self.case_number)
        self.query.exec()

    def load_data_into_case(self) -> None:
        while self.query.next():
            if self.case.case_number is None:
                self.load_case_information()
            self.load_charge_information()

    def load_case_information(self) -> None:
        self.case.case_number = self.query.value(1)
        self.case.defendant.last_name = self.query.value(2).title()
        self.case.defendant.first_name = self.query.value(3).title()
        self.case.fra_in_file = self.query.value(7)
        self.case.defense_counsel = f"{self.query.value(10).title()} {self.query.value(9).title()}"
        self.case.defense_counsel_type = self.query.value(11)

    def load_charge_information(self) -> None:
        offense = self.clean_offense_name(self.query.value(4))
        statute = self.clean_statute_name(self.query.value(5))
        degree = self.query.value(6)
        moving_bool = self.query.value(8)
        charge = (offense, statute, degree, moving_bool)
        self.case.charges_list.append(charge)

    def clean_statute_name(self, statute: str) -> str:
        """Removes trailing asteriks that are part of CMS data."""
        return statute.rstrip("*")

    def clean_offense_name(self, offense: str) -> str:
        """Sets an offense name to title case, but leaves certain abbreviations and removes
        degree of charge if in offense name."""
        offense_word_list = offense.split()
        clean_offense_word_list = []
        for word in offense_word_list:
            if word == "OMVI":
                clean_offense_word_list.append("OVI")
                continue
            elif word == "FRA/JUDGMENT":
                clean_offense_word_list.append("FRA / Judgment")
                continue
            elif word in self.abbreviation_list:
                clean_offense_word_list.append(word)
                continue
            elif word in self.delete_word_list:
                continue
            else:
                clean_offense_word_list.append(word.capitalize())
        clean_offense = " ".join([str(word) for word in clean_offense_word_list])
        return clean_offense

    def load_case(self) -> CmsCaseInformation:
        return self.case


def open_db_connection(connection_name: str) -> QSqlDatabase:
    db_connection = QSqlDatabase.database(connection_name, open=True)
    check_if_db_open(db_connection, connection_name)
    logger.log('DATABASE', f'{db_connection.connectionName()} database connection open.')
    return db_connection


def close_db_connection(db_connection: QSqlDatabase) -> None:
    db_connection.close()
    logger.log('DATABASE', f'{db_connection.connectionName()} database connection closed.')


def remove_db_connection(connection_name: str) -> None:
    QSqlDatabase.removeDatabase(connection_name)
    logger.log('DATABASE', f'{connection_name} database connection removed.')


def create_db_connection(database_name: str, connection_name: str) -> QSqlDatabase:
    if not os.path.exists(database_name):
        logger.warning("The database does not exist. Creating new database.")
    db_connection = create_db(database_name, connection_name)
    return db_connection


def create_db(database_name: str, connection_name: str) -> QSqlDatabase:
    db_connection = QSqlDatabase.addDatabase("QSQLITE", connection_name)
    db_connection.setDatabaseName(database_name)
    return db_connection


def check_if_db_open(db_connection: QSqlDatabase, connection_name: str) -> bool:
    if not db_connection.isOpen():
        logger.critical(f"Unable to connect to {connection_name} database")
        sys.exit(1)
    return True


def create_daily_case_list_sql_tables(con_daily_case_lists: QSqlDatabase) -> None:
    for item in EXCEL_DAILY_CASE_LISTS:
        table_name: str
        table_name = item[1]
        QSqlQuery(con_daily_case_lists).exec(create_daily_case_list_tables_sql_query(table_name))


def load_daily_case_list_data(con_daily_case_lists: QSqlDatabase) -> None:
    for item in EXCEL_DAILY_CASE_LISTS:
        excel_report: str
        table_name: str
        excel_report, table_name = item
        delete_existing_sql_table(con_daily_case_lists, table_name)
        insert_daily_case_list_sql_data(con_daily_case_lists, excel_report, table_name)


def insert_daily_case_list_sql_data(
    con_daily_case_lists: QSqlDatabase, excel_report: str, table_name: str
) -> None:
    cases_from_table = return_cases_data_from_excel(f"{DB_PATH}{excel_report}")
    insert_data_query = QSqlQuery(con_daily_case_lists)
    insert_data_query.prepare(insert_daily_case_list_tables_sql_query(table_name))
    for case in cases_from_table:
        insert_data_query.addBindValue(case.case_number)
        insert_data_query.addBindValue(case.defendant_last_name)
        insert_data_query.addBindValue(case.defendant_first_name)
        insert_data_query.addBindValue(case.offense)
        insert_data_query.addBindValue(case.statute)
        insert_data_query.addBindValue(case.degree)
        insert_data_query.addBindValue(case.fra_in_file)
        insert_data_query.addBindValue(case.moving_bool)
        insert_data_query.addBindValue(case.def_atty_last_name)
        insert_data_query.addBindValue(case.def_atty_first_name)
        insert_data_query.addBindValue(case.def_atty_type)
        insert_data_query.exec()


def delete_existing_sql_table(db_connection: QSqlDatabase, table_name: str) -> None:
    delete_table = QSqlQuery(db_connection)
    delete_table.prepare(delete_table_sql_query(table_name))
    delete_table.exec()


def query_offense_statute_data(db_connection: QSqlDatabase, query_value: str) -> list:
    query_string = select_distinct_offense_statute_sql_query()
    query = QSqlQuery(db_connection)
    query.prepare(query_string)
    query.exec()
    item_list = []
    while query.next():
        if query_value == "offense":
            item_list.append(query.value(0))
        elif query_value == "statute":
            item_list.append(query.value(1))
    item_list.sort()
    return item_list


def query_daily_case_list_data(table: str, db_connection: QSqlDatabase) -> list:
    query_string = select_distinct_def_last_def_first_case_number_sql_query(table)
    query = QSqlQuery(db_connection)
    query.prepare(query_string)
    query.exec()
    item_list = []
    while query.next():
        last_name = query.value(0).title()
        case_number = query.value(2)
        name = f"{last_name} - {case_number}"
        item_list.append(name)
    item_list.sort()
    item_list.insert(0, "")
    return item_list


def create_charges_sql_table(con_charges: str) -> None:
    QSqlQuery(con_charges).exec(create_charges_table_sql_query())


def insert_charges_sql_data(con_charges: QSqlDatabase, table_name: str) -> None:
    charges_from_table = create_charges_data_list(CHARGES_TABLE)
    for charge in charges_from_table:
        insert_data_query = QSqlQuery(con_charges)
        insert_data_query.prepare(insert_charges_sql_query(table_name, charge))
        insert_data_query.exec()


def load_charges_data(con_charges: QSqlDatabase) -> None:
    delete_existing_sql_table(con_charges, "charges")
    insert_charges_sql_data(con_charges, "charges")
    con_charges.close()


def sql_query_offense_type(key: str) -> str:
    """This is called from the AddChargeDialogSlotFunctions to set the offense type to calculate
    court costs. If no match is found this returns "Moving" which is the highest court cost, so if
    the defendant is told the amount it should not be less than what it is owed.
    TODO: add for AmendChargeDialogSlotFunctions."""
    charges_database = open_db_connection("con_charges")
    query = QSqlQuery(charges_database)
    query.prepare(select_statute_from_charges_for_offense_type_sql_query())
    query.bindValue(":key", key)
    query.exec()
    offense_type = "Moving"
    while query.next():
        statute = query.value(2)
        offense_type = query.value(4)
        if statute == key:
            query.finish()
            charges_database.close()
            return offense_type
    return offense_type


def extract_data(case_data: dict) -> None:
    wb_name = "Case_Data.xlsx"
    wb_name = DB_PATH + wb_name
    wb = load_workbook(wb_name)
    page = wb.active
    case_number = case_data.get("case_number")
    judicial_officer = case_data.get("judicial_officer").last_name
    charges_list = case_data.get("charges_list")
    max_row = page.max_row
    max_row = max_row + 1
    for index, charge in enumerate(charges_list):
        page.cell(row=max_row + index, column=1, value=case_number)
        page.cell(row=max_row + index, column=2, value=judicial_officer)
        page.cell(row=max_row + index, column=3, value=charge.get("offense"))
        page.cell(row=max_row + index, column=4, value=charge.get("statute"))
        page.cell(row=max_row + index, column=5, value=charge.get("degree"))
        page.cell(row=max_row + index, column=6, value=charge.get("plea"))
        page.cell(row=max_row + index, column=7, value=charge.get("finding"))
        page.cell(row=max_row + index, column=8, value=charge.get("fines_amount"))
        page.cell(row=max_row + index, column=9, value=charge.get("fines_suspended"))
        page.cell(row=max_row + index, column=10, value=charge.get("jail_days"))
        page.cell(row=max_row + index, column=11, value=charge.get("jail_days_suspended"))
    try:
        wb.save(filename=wb_name)
    except PermissionError:
        pass


def main():

    create_db_connection(f"{DB_PATH}daily_case_lists.sqlite", "con_daily_case_lists")
    con_daily_case_lists = open_db_connection("con_daily_case_lists")
    create_daily_case_list_sql_tables(con_daily_case_lists)
    load_daily_case_list_data(con_daily_case_lists)
    close_db_connection(con_daily_case_lists)

    create_db_connection(f"{DB_PATH}charges.sqlite", "con_charges")
    con_charges = open_db_connection("con_charges")
    create_charges_sql_table(con_charges)
    load_charges_data(con_charges)
    close_db_connection(con_charges)


if __name__ == "__main__":
    logger.log('IMPORT', f'{__name__} run directly.')
else:
    main()
    logger.log('IMPORT', f'{__name__} imported.')
