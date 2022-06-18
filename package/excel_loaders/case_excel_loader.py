"""Module for loading case data from Excel files."""

from openpyxl import Workbook  # type: ignore

from package.excel_loaders.loader_functions import (
    create_headers_dict,
    get_excel_file_headers,
    load_active_worksheet,
)
from package.models.excel_models import CaseExcelData

NO_DATA = 'No Data'
COL_CASE = 'Case'
COL_DEF_LAST_NAME = 'Lastname'
COL_DEF_FIRST_NAME = 'Firstname'
COL_CHARGE = 'Charge'
COL_STATUTE = 'Code'
COL_DEGREE = 'Degree'
COL_INSURANCE = 'Insurance'
COL_MOVING_OFFENSE = 'Moving'
COL_DEF_ATTY_LAST_NAME = 'Atty Last'
COL_DEF_ATTY_FIRST_NAME = 'Atty First'
COL_DEF_ATTY_TYPE = 'Atty Type'


def create_case_data_list(ws: Workbook.active, headers_dict: dict) -> list[CaseExcelData]:
    """Returns a list of CaseExcelData objects.

    The CaseExcelData object is loaded with the data from each row (case) and the attributes
    of each case are mapped to the specific column that contains the case attribute. The load
    portion of the function starts with row 2 because row 1 is headers.
    """
    case_data_list: list = []
    row_count: int = ws.max_row + 1
    for row in range(2, row_count):
        case = CaseExcelData()
        case.case_number = get_cell_value(ws, row, headers_dict[COL_CASE])
        case.defendant_last_name = get_cell_value(ws, row, headers_dict[COL_DEF_LAST_NAME])
        case.defendant_first_name = get_cell_value(ws, row, headers_dict[COL_DEF_FIRST_NAME])
        case.offense = get_cell_value(ws, row, headers_dict[COL_CHARGE])
        case.statute = get_cell_value(ws, row, headers_dict[COL_STATUTE])
        case.degree = get_cell_value(ws, row, headers_dict[COL_DEGREE])
        case.fra_in_file = get_cell_value(ws, row, headers_dict[COL_INSURANCE])
        case.moving_bool = get_cell_value(ws, row, headers_dict[COL_MOVING_OFFENSE])
        case.def_atty_last_name = get_cell_value(ws, row, headers_dict[COL_DEF_ATTY_LAST_NAME])
        case.def_atty_first_name = get_cell_value(ws, row, headers_dict[COL_DEF_ATTY_FIRST_NAME])
        case.def_atty_type = get_cell_value(ws, row, headers_dict[COL_DEF_ATTY_TYPE])
        case_data_list.append(case)
    return case_data_list


def return_cases_data_from_excel(excel_file: str) -> list[CaseExcelData]:
    """Loads active worksheet and generates header dict in order to create case data list."""
    ws = load_active_worksheet(excel_file)
    header_list = get_excel_file_headers(ws)
    headers_dict = create_headers_dict(header_list)
    return create_case_data_list(ws, headers_dict)


def get_cell_value(ws: Workbook.active, row: int, col: int) -> str:
    """Returns the cell value for a cell in the active excel worksheet."""
    if ws.cell(row=row, column=col).value is None:
        return set_cell_value_if_none(ws, col)
    return ws.cell(row=row, column=col).value


def set_cell_value_if_none(ws: Workbook.active, col: int) -> str:
    """Returns a specific string for certain attributes if there is no data."""
    if ws.cell(row=1, column=col).value == COL_INSURANCE:
        return 'U'
    if ws.cell(row=1, column=col).value == COL_DEF_ATTY_LAST_NAME:
        return ''
    if ws.cell(row=1, column=col).value == COL_DEF_ATTY_FIRST_NAME:
        return ''
    return NO_DATA
