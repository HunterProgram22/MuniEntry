"""Module that contains general functions used for loading data from Excel files."""
from openpyxl import Workbook, load_workbook  # type: ignore


def get_excel_file_headers(worksheet: Workbook.active) -> list[tuple[str, int]]:
    """Returns a list of tuples from first row in excel file.

    Values are packaged as a key (col.value = cell text) and value (col.column =
    cell column) pair from the first row of the active worksheet in excel file.
    """
    header_list = []
    for col in next(worksheet.iter_rows(min_row=1, max_row=1)):
        column_info = (col.value, col.column)
        header_list.append(column_info)
    return header_list


def load_active_worksheet(excel_file: str) -> Workbook.active:
    """Returns the active worksheet from the excel file that is loaded."""
    workbook = load_workbook(excel_file)
    return workbook.active


def create_headers_dict(headers_list: list[tuple[str, int]]) -> dict:
    """Returns a dict with the column name (key) mapped to the column (value).

    The headers_list is a tuple of the column name and column from row 1 of the
    excel file.
    """
    headers_dict = {}
    for col_tuple in headers_list:
        col_name, col_value = col_tuple
        headers_dict[col_name] = col_value
    return headers_dict
