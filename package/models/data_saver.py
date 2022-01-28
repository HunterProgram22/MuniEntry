import pathlib

from openpyxl import load_workbook

from settings import DB_PATH


def extract_data(case_data):
    wb_name = "Case_Data.xlsx"
    wb_name = DB_PATH + wb_name
    wb = load_workbook(wb_name)
    page = wb.active
    case_number = case_data.get('case_number')
    judicial_officer = case_data.get('judicial_officer').last_name
    charges_list = case_data.get('charges_list')
    max = page.max_row
    max = max + 1
    for index, charge in enumerate(charges_list):
        page.cell(row=max+index, column=1, value=case_number)
        page.cell(row=max+index, column=2, value=judicial_officer)
        page.cell(row=max+index, column=3, value=charge.get('offense'))
        page.cell(row=max+index, column=4, value=charge.get('statute'))
        page.cell(row=max+index, column=5, value=charge.get('degree'))
        page.cell(row=max+index, column=6, value=charge.get('plea'))
        page.cell(row=max+index, column=7, value=charge.get('finding'))
        page.cell(row=max+index, column=8, value=charge.get('fines_amount'))
        page.cell(row=max+index, column=9, value=charge.get('fines_suspended'))
        page.cell(row=max+index, column=10, value=charge.get('jail_days'))
        page.cell(row=max+index, column=11, value=charge.get('jail_days_suspended'))
    try:
        wb.save(filename=wb_name)
    except PermissionError:
        pass

