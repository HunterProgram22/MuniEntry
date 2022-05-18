import re
import datetime

from PyQt5 import QtCore
from PyQt5.QtCore import QSortFilterProxyModel, Qt, QEvent, QDate, QDateTime, QTime
from PyQt5.QtGui import QIntValidator, QFont
from PyQt5.QtWidgets import QPushButton, QMessageBox, QComboBox, QLineEdit, QCheckBox, QCompleter, \
    QInputDialog, QDateEdit, QTimeEdit, QMenu, QLabel, QDateTimeEdit
from PyQt5 import QtGui
from openpyxl import load_workbook  # type: ignore

from settings import ICON_PATH, DB_PATH


def return_attorney_data_from_excel(excel_file: str) -> list:
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    max_row = worksheet.max_row
    max_row = max_row + 1
    for row in range(2, max_row):
        attorney_first_name = worksheet.cell(row=row, column=1)
        attorney_last_name = worksheet.cell(row=row, column=2)
        attorney_full_name = f"{attorney_first_name.value} {attorney_last_name.value}"
        data.append(attorney_full_name)
    return data

TODAY = QtCore.QDate.currentDate()
TODAY_STRING = TODAY.toString("MMMM dd, yyyy")
ATTORNEY_LIST = return_attorney_data_from_excel(f"{DB_PATH}Attorneys.xlsx")


class NoScrollComboBox(QComboBox):
    def __init__(self, parent=None):
        super(QComboBox, self).__init__(parent)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)

    def wheelEvent(self, event):
        if event == QtCore.QEvent.Wheel:
            event.ignore()


class NoScrollDateEdit(QDateEdit):
    def __init__(self, parent=None):
        super(QDateEdit, self).__init__(parent)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setDate(TODAY)

    def get_date(self) -> str:
        return self.date().toString("MMMM dd, yyyy")

    def set_date(self, date_str: str) -> QDate:
        if date_str is None:
            date_str = TODAY_STRING
        date_str = QDate.fromString(date_str, "MMMM dd, yyyy")
        return self.setDate(date_str,)

    def wheelEvent(self, event):
        if event == QtCore.QEvent.Wheel:
            event.ignore()


class NoScrollTimeEdit(QTimeEdit):
    def __init__(self, parent=None):
        super(QTimeEdit, self).__init__(parent)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)

    def get_time(self) -> str:
        return self.time().toString("hh:mm A")

    def set_time(self, time_str: str) -> QTime:
        if time_str is None:
            time_str = "08:30 AM"
        time_str = QTime.fromString(time_str, "hh:mm A")
        return self.setTime(time_str)


    def wheelEvent(self, event):
        if event == QtCore.QEvent.Wheel:
            event.ignore()


class ExtendedComboBox(QComboBox):
    def __init__(self, parent=None):
        """Code is from the user Joachim Bonfert 8/2/21 answer at https://stackoverflow.com/questions/4827207/
        how-do-i-filter-the-pyqt-qcombobox-items-based-on-the-text-input"""
        super(ExtendedComboBox, self).__init__(parent)
        self.setFocusPolicy(Qt.StrongFocus)
        self.setEditable(True)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_menu)

        # add a filter model to filter matching items
        self.pFilterModel = QSortFilterProxyModel(self)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.model())

        # add a completer, which uses the filter model
        self.completer = QCompleter(self.pFilterModel, self)
        # always show all (filtered) completions
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.setCompleter(self.completer)

        # connect signals
        self.lineEdit().textEdited[str].connect(self.pFilterModel.setFilterFixedString)
        self.completer.activated.connect(self.on_completer_activated)

    def show_menu(self, pos):
        menu = QMenu()
        menu.addAction("Delete Case", self.delete_case)
        menu.exec_(self.mapToGlobal(pos))

    def delete_case(self):
        index = self.currentIndex()
        self.removeItem(index)

    def keyPressEvent(self, event: QtGui.QKeyEvent) -> None:
        super(ExtendedComboBox, self).keyPressEvent(event)
        if event.key() == QtCore.Qt.Key_Delete:
            self.delete_case()

    # on selection of an item from the completer, select the corresponding item from combobox
    def on_completer_activated(self, text):
        if text:
            index = self.findText(text)
            self.setCurrentIndex(index)
            self.activated[str].emit(self.itemText(index))

    # on model change, update the models of the filter and completer as well
    def setModel(self, model):
        super(ExtendedComboBox, self).setModel(model)
        self.pFilterModel.setSourceModel(model)
        self.completer.setModel(self.pFilterModel)

    # on model column change, update the model column of the filter and completer as well
    def setModelColumn(self, column):
        self.completer.setCompletionColumn(column)
        self.pFilterModel.setFilterKeyColumn(column)
        super(ExtendedComboBox, self).setModelColumn(column)


class DefenseCounselComboBox(NoScrollComboBox):
    def __init__(self, parent=None):
        super(NoScrollComboBox, self).__init__(parent)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)

    def load_attorneys(self):
        for attorney in ATTORNEY_LIST:
            self.addItem(attorney)


class StatuteLineEdit(QLabel):
    def __init__(self, statute: str=None, parent=None):
        super(QLabel, self).__init__(parent)
        self.setStyleSheet("font: bold 'Palatino Linotype';"
                           "font-size: 10pt")
        self.set_up_widget(statute)
        self.statute = statute

    def get_text(self) -> str:
        """This method exists to return the string of the statute that is used in the entry
        as opposed to the text that is displayed which is a hyperlink."""
        return self.statute

    def set_up_widget(self, statute: str):
        self.statute = statute
        self.setMinimumSize(QtCore.QSize(200, 0))
        self.setMaximumSize(QtCore.QSize(200, 50))
        self.setObjectName("statute_lineEdit")
        url_link = self.set_url_link(self.statute)
        self.setText(url_link)
        self.setOpenExternalLinks(True)

    def set_url_link(self, statute: str) -> str:
        admin_code = "\d\d\d\d.\d\d-\d-\d\d"
        url_statute = re.search(admin_code, statute)
        if url_statute is not None:
            url_statute = url_statute.group()
            url_statute = url_statute.replace('.', ':')
            return f"<a href=\'https://codes.ohio.gov/ohio-administrative-code/rule-{url_statute}\'>{statute}</a>"
        admin_code_two = "\d\d\d\d.\d\d-\d\d-\d\d"
        url_statute = re.search(admin_code_two, statute)
        if url_statute is not None:
            url_statute = url_statute.group()
            url_statute = url_statute.replace('.', ':')
            return f"<a href=\'https://codes.ohio.gov/ohio-administrative-code/rule-{url_statute}\'>{statute}</a>"
        seven_digit_stat = "\d\d\d\d.\d\d\d"
        url_statute = re.search(seven_digit_stat, statute)
        if url_statute is not None:
            url_statute = url_statute.group()
            return f"<a href=\'https://codes.ohio.gov/ohio-revised-code/section-{url_statute}\'>{statute}</a>"
        six_digit_stat = "\d\d\d\d.\d\d"
        url_statute = re.search(six_digit_stat, statute)
        if url_statute is not None:
            url_statute = url_statute.group()
            return f"<a href=\'https://codes.ohio.gov/ohio-revised-code/section-{url_statute}\'>{statute}</a>"
        five_digit_stat = "\d\d\d.\d\d"
        url_statute = re.search(five_digit_stat, statute)
        if url_statute is not None:
            url_statute = url_statute.group()
            return f"<a href=\'https://library.municode.com/oh\'>{statute}</a>"
        return f"<a href=\'https://www.google.com/\'>{statute}</a>"


class DegreeComboBox(NoScrollComboBox):
    def __init__(self, degree, parent=None):
        super(NoScrollComboBox, self).__init__(parent)
        self.set_up_widget(degree)

    def set_up_widget(self, degree):
        self.setMinimumSize(QtCore.QSize(200, 0))
        self.setMaximumSize(QtCore.QSize(200, 50))
        self.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.setEditable(False)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setObjectName("degree_choice_box")
        self.addItem("")
        self.addItem("M1")
        self.addItem("M2")
        self.addItem("M3")
        self.addItem("M4")
        self.addItem("MM")
        self.addItem("UCM")
        self.addItem("No Data")
        self.setCurrentText(degree)


class PleaComboBox(NoScrollComboBox):
    def __init__(self, column, parent=None):
        super(NoScrollComboBox, self).__init__(parent)
        self.column = column
        self.set_up_widget()

    def set_up_widget(self):
        self.setMinimumSize(QtCore.QSize(200, 0))
        self.setMaximumSize(QtCore.QSize(200, 50))
        self.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.setEditable(False)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setObjectName("plea_choice_box")
        self.addItem("")
        self.addItem("Guilty")
        self.addItem("No Contest")
        self.addItem("Not Guilty")
        self.addItem("Dismissed")


class FindingComboBox(NoScrollComboBox):
    def __init__(self, parent=None):
        super(NoScrollComboBox, self).__init__(parent)
        self.set_up_widget()

    def set_up_widget(self):
        self.setMinimumSize(QtCore.QSize(200, 0))
        self.setMaximumSize(QtCore.QSize(200, 50))
        self.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.setEditable(False)
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.setObjectName("finding_choice_box")
        self.addItem("")
        self.addItem("Guilty")
        self.addItem("Not Guilty")
        self.addItem("Lesser Included")
        self.addItem("Guilty - Allied Offense")
        self.addItem("Not Guilty - Allied Offense")


class ChargeGridDeleteButton(QPushButton):
    def __init__(self, column_index, charge, dialog, parent=None):
        super(QPushButton, self).__init__(parent)
        self.column_index = column_index
        self.dialog = dialog
        self.charge = charge
        self.set_up_widget()

    def set_up_widget(self):
        self.setStyleSheet("background-color: rgb(170, 58, 63);")
        self.setText("Delete")
        self.setObjectName("charge_grid_delete_Button")
        self.setFocusPolicy(QtCore.Qt.NoFocus)
        self.pressed.connect(self.delete_charge_from_grid_and_charges_list)

    def delete_charge_from_grid_and_charges_list(self):
        """Uses the delete_button that is indexed to the column to delete the
        QLabels for the charge."""
        self.dialog.entry_case_information.charges_list.remove(self.charge)
        for row in range(self.dialog.charges_gridLayout.rowCount()):
            layout_item = self.dialog.charges_gridLayout.itemAtPosition(row, self.column_index)
            if layout_item is not None:
                layout_item.widget().deleteLater()
                self.dialog.charges_gridLayout.removeItem(layout_item)


class ChargeGridAmendButton(QPushButton):
    def __init__(self, column_index, charge, dialog, parent=None):
        super(QPushButton, self).__init__(parent)
        self.column_index = column_index
        self.dialog = dialog
        self.charge = charge
        self.set_up_widget()

    def set_up_widget(self):
        self.setStyleSheet("background-color: rgb(62, 146, 255);")
        self.setText("Amend")
        self.setObjectName("charge_grid_amend_Button")
        self.setFocusPolicy(QtCore.Qt.NoFocus)
        self.released.connect(self.dialog.functions.start_amend_offense_dialog)


class AlliedCheckbox(QCheckBox):
    def __init__(self, column_index, dialog, parent=None):
        super(QCheckBox, self).__init__(parent)
        self.column_index = column_index
        self.dialog = dialog
        self.set_up_widget()

    def set_up_widget(self):
        self.setText("Allied Offense")
        self.setObjectName("allied_checkBox")
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.toggled.connect(self.set_to_allied)

    def set_to_allied(self):
        if self.isChecked():
            try:
                if self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().currentText() == "Guilty":
                    self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setCurrentText("Guilty - Allied Offense")
                elif self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().currentText() == "Not Guilty":
                    self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setCurrentText("Not Guilty - Allied Offense")
            except AttributeError:
                pass
        else:
            try:
                if self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().currentText() == "Guilty - Allied Offense":
                    self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setCurrentText(
                        "Guilty")
                elif self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().currentText() == "Not Guilty - Allied Offense":
                    self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setCurrentText(
                        "Not Guilty")
            except AttributeError:
                pass


class DismissedCheckbox(QCheckBox):
    def __init__(self, column_index, dialog, parent=None):
        super(QCheckBox, self).__init__(parent)
        self.dialog = dialog
        self.column_index = column_index
        self.set_up_widget()

    def set_up_widget(self):
        self.setText("Offense Dismissed")
        self.setObjectName("dismissed_checkBox")
        self.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.toggled.connect(self.set_to_dismissed)

    def set_to_dismissed(self):
        """TODO: the try except block is to account for the Leap Dialogs and NoJail not having same # rows. Fix."""
        if self.isChecked():
            try:
                self.dialog.charges_gridLayout.itemAtPosition(5, self.column_index).widget().setCurrentText("Dismissed")
            except AttributeError:
                self.dialog.charges_gridLayout.itemAtPosition(4, self.column_index).widget().setCurrentText("Dismissed")
                self.dialog.charges_gridLayout.itemAtPosition(5, self.column_index).widget().setHidden(True)
            try:
                self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(7, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(8, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(9, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(10, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(11, self.column_index).widget().setHidden(True)
                self.dialog.charges_gridLayout.itemAtPosition(12, self.column_index).widget().setHidden(True)
            except AttributeError:
                pass
        else:
            try:
                self.dialog.charges_gridLayout.itemAtPosition(5, self.column_index).widget().setCurrentText("")
            except AttributeError:
                self.dialog.charges_gridLayout.itemAtPosition(4, self.column_index).widget().setCurrentText("")
                self.dialog.charges_gridLayout.itemAtPosition(5, self.column_index).widget().setHidden(False)
            try:
                self.dialog.charges_gridLayout.itemAtPosition(6, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(7, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(8, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(9, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(10, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(11, self.column_index).widget().setHidden(False)
                self.dialog.charges_gridLayout.itemAtPosition(12, self.column_index).widget().setHidden(False)
            except AttributeError:
                pass


class RequiredBox(QMessageBox):
    def __init__(self, message, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.message = message
        self.set_up_widget()

    def set_up_widget(self):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        self.setIcon(QMessageBox.Critical)
        self.setWindowTitle("Required")
        self.setText(self.message)
        self.setStandardButtons(QMessageBox.Ok)


class InfoBox(QMessageBox):
    def __init__(self, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.set_up_widget()

    def set_up_widget(self):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        self.setIcon(QMessageBox.Information)
        self.setStandardButtons(QMessageBox.Ok)


class WarningBox(QMessageBox):
    def __init__(self, message, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.message = message
        self.set_up_widget()

    def set_up_widget(self):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        self.setIcon(QMessageBox.Warning)
        self.setWindowTitle("Warning")
        self.setText(self.message)
        self.setStandardButtons(QMessageBox.Yes | QMessageBox.No)


class TwoChoiceQuestionBox(QMessageBox):
    def __init__(self, message, yes_choice, no_choice, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.message = message
        self.set_up_widget(yes_choice, no_choice)

    def set_up_widget(self, yes_choice, no_choice):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        self.setIcon(QMessageBox.Question)
        self.setWindowTitle("Additional Information Required")
        self.setText(self.message)
        self.addButton(QPushButton(yes_choice), QMessageBox.YesRole) # YesRole returns 5
        self.addButton(QPushButton(no_choice), QMessageBox.NoRole)  # NoRole returns 6


class DataInputBox(QInputDialog):
    def __init__(self, message, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.message = message
        self.set_up_widget()

    def set_up_widget(self):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        layout = QFormLayout()
        self.box = QLineEdit()


class JailWarningBox(QMessageBox):
    def __init__(self, message, parent=None):
        super(QMessageBox, self).__init__(parent)
        self.message = message
        self.set_up_widget()

    def set_up_widget(self):
        self.setWindowIcon(QtGui.QIcon(ICON_PATH + 'gavel.ico'))
        self.setIcon(QMessageBox.Warning)
        self.setWindowTitle("Warning - No Jail Report Date Set")
        self.setText(self.message)
        self.setStandardButtons(QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)


# DataValidator Widgets
class IntegerValidator(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.validator = QIntValidator(0, 1000, self)
        self.setValidator(self.validator)


class ChargeGridIntegerWidget(IntegerValidator):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.set_up_widget()

    def set_up_widget(self):
        self.setMinimumSize(QtCore.QSize(200, 0))
        self.setMaximumSize(QtCore.QSize(200, 50))
        self.setStyleSheet("background-color: rgb(255, 255, 255);")


class FineLineEdit(ChargeGridIntegerWidget):
    def __init__(self, offense=None, parent=None):
        super().__init__(parent)
        self.setObjectName("fines_amount")
        self.set_fine_amount(offense)

    def set_fine_amount(self, offense):
        if offense == "Seatbelt - Driver":
            self.setText("30")
        elif offense == "Seatbelt - Passenger":
            self.setText("20")
        elif offense == "Failure to Stop for School Bus":
            self.setText("500")
        else:
            self.setText("")


class FineSuspendedLineEdit(ChargeGridIntegerWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("fines_suspended")


class JailLineEdit(ChargeGridIntegerWidget):
    def __init__(self, offense=None, parent=None):
        super().__init__(parent)
        self.setObjectName("jail_days")


class JailSuspendedLineEdit(ChargeGridIntegerWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("jail_days_suspended")


class JailTimeCreditLineEdit(ChargeGridIntegerWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("jail_time_credit_box")


def main():
    pass

if __name__ == "__main__":
   # stuff only to run when not called via 'import' here
   main()
