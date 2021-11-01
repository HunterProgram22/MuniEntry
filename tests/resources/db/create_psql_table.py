"""This module is used to create the charges table.

TODO: Right now it manually needs the rows in range changed
when charges are added, need to update to call len or something
similar.

TODO: Also, eventually should this be called on load so charges
could be added somewhere by anyone and the db created each time
the application is loaded to account for new charges. May be
a bit unnecessary - but perhaps adding charges by user to a shared
db could be done."""
import sys
import pathlib
from loguru import logger

from openpyxl import load_workbook

from PyQt5.QtSql import QSqlDatabase, QSqlQuery


PATH = str(pathlib.Path().absolute())

EXCEL_FILE = PATH + "\Case_Types.xlsx"



@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    for row in range(2, 35):
        offense = worksheet.cell(row=row, column=1)
        statute = worksheet.cell(row=row, column=2)
        degree = worksheet.cell(row=row, column=3)
        type = worksheet.cell(row=row, column=4)
        charge = (offense.value, statute.value, degree.value, type.value)
        data.append(charge)
    return data

# Connecting both an alphabetical (offense) and numerical (statute) ordered database
con = QSqlDatabase.addDatabase("QSQLITE")
con.setDatabaseName(PATH + "\\charges.sqlite")


if not con.open():
    print("Unable to connect to database")
    sys.exit(1)


# Create a query and execute it right away using .exec()
createTableQuery = QSqlQuery()
createTableQuery.exec(
    """
    CREATE TABLE charges (
        id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
        offense VARCHAR(60) NOT NULL,
        statute VARCHAR(50) NOT NULL,
        degree VARCHAR(50) NOT NULL,
        type VARCHAR(50) NOT NULL
    )
    """
)

insertDataQuery = QSqlQuery()
insertDataQuery.prepare(
    """
    INSERT INTO charges (
        offense,
        statute,
        degree,
        type
    )
    VALUES (?, ?, ?, ?)
    """
)


# TO POPULATE A COMBO BOX http://www.voidynullness.net/blog/2013/02/05/qt-populate-combo-box-from-database-table/
# https://python-forum.io/thread-11659.html
# Create two tables one for alpha sort and one for num sort - defintely a better way to do this
data_from_table = return_data_from_excel(EXCEL_FILE)
print(data_from_table)

# Use .addBindValue() to insert data
for offense, statute, degree, type in data_from_table:
    insertDataQuery.addBindValue(offense)
    insertDataQuery.addBindValue(statute)
    insertDataQuery.addBindValue(degree)
    insertDataQuery.addBindValue(type)
    insertDataQuery.exec()
