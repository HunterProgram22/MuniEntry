import sys
import pathlib
from loguru import logger

from openpyxl import load_workbook


PATH = str(pathlib.Path().absolute())

EXCEL_FILE = PATH + "\Case_Types.xlsx"

data = []

@logger.catch
def return_data_from_excel(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    for row in range(2, 30):
        offense = worksheet.cell(row=row, column=1)
        statute = worksheet.cell(row=row, column=2)
        degree = worksheet.cell(row=row, column=3)
        charge = (offense.value, statute.value, degree.value)
        data.append(charge)



return_data_from_excel(EXCEL_FILE)
print(data)
