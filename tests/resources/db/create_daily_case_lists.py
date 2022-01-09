"""This module is used to create the slated table."""
import sys
import os
import pathlib
from loguru import logger

from openpyxl import load_workbook

from PyQt5.QtSql import QSqlDatabase, QSqlQuery

"""This path is from the MuniEntry level of the program because it uses Path at the level of the module
into which it is imported."""
PATH = str(pathlib.Path().absolute())
DB_PATH = PATH + "\\resources\\db\\"


@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    row_count = worksheet.max_row + 1
    for row in range(2,row_count):
        case_number = worksheet.cell(row=row, column=1)
        defendant_last_name = worksheet.cell(row=row, column=3)
        defendant_first_name = worksheet.cell(row=row, column=4)
        offense = worksheet.cell(row=row, column=5)
        statute = worksheet.cell(row=row, column=6)
        degree = worksheet.cell(row=row, column=7)
        if worksheet.cell(row=row, column=8).value == None:
            worksheet.cell(row=row, column=8).value = "U"
            fra_in_file = worksheet.cell(row=row, column=8)
        else:
            fra_in_file = worksheet.cell(row=row, column=8)
        case = (case_number.value,
                defendant_last_name.value,
                defendant_first_name.value,
                offense.value,
                statute.value,
                degree.value,
                fra_in_file.value,
                )
        data.append(case)
    return data

def main():
    """database_list contains tuples for the items to create each database for the daily case
    lists. Tuple format is (Excel_file_to_load, database_path_name, database_connection_name)."""
    database_list = [
        ("Slated.xlsx", "slated.sqlite", "slated_table"),
        ("Arraignments.xlsx", "arraignments.sqlite", "arraignments_table"),
        ("Final_Pretrials.xlsx", "final_pretrials.sqlite", "final_pretrials_table"),
    ]
    for items in database_list:
        if os.path.exists(DB_PATH + items[1]):
            os.remove(DB_PATH + items[1])
        else:
            print("The file does not exist")
        con = QSqlDatabase.addDatabase("QSQLITE", items[2])
        con.setDatabaseName(DB_PATH + items[1])
        if not con.open():
            print("Unable to connect to database")
            sys.exit(1)
        # Create a query and execute it right away using .exec()
        createTableQuery = QSqlQuery(con)
        createTableQuery.exec(
            """
            CREATE TABLE cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
                case_number VARCHAR(20) NOT NULL,
                defendant_last_name VARCHAR(50) NOT NULL,
                defendant_first_name VARCHAR(50) NOT NULL,
                offense VARCHAR(80) NOT NULL,
                statute VARCHAR(50) NOT NULL,
                degree VARCHAR(10) NOT NULL,
                fra_in_file VARCHAR(5) NOT NULL
            )
            """
        )
        insertDataQuery = QSqlQuery(con)
        # Do not add comma to last value inserted
        insertDataQuery.prepare(
            """
            INSERT INTO cases (
                case_number,
                defendant_last_name,
                defendant_first_name,
                offense,
                statute,
                degree,
                fra_in_file
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """
        )
        data_from_table = return_data_from_excel(DB_PATH + items[0])
        # Use .addBindValue() to insert data
        for case_number, defendant_last_name, defendant_first_name, offense, statute, degree, fra_in_file in data_from_table:
            insertDataQuery.addBindValue(case_number)
            insertDataQuery.addBindValue(defendant_last_name)
            insertDataQuery.addBindValue(defendant_first_name)
            insertDataQuery.addBindValue(offense)
            insertDataQuery.addBindValue(statute)
            insertDataQuery.addBindValue(degree)
            insertDataQuery.addBindValue(fra_in_file)
            insertDataQuery.exec()
        con.close()
        con.removeDatabase("QSQLITE")
    return None

if __name__ == "__main__":
    print("Daily Case Lists created directly from script")
else:
    main()
    print("Imported Daily Case List Tables")
