import os
import sqlite3
import sys
from abc import ABC, abstractmethod
import string

from openpyxl import load_workbook  # type: ignore
from PyQt5.QtSql import QSqlQuery, QSqlDatabase

from db.sql_queries import (
    create_daily_case_list_tables_sql_query,
    insert_daily_case_list_tables_sql_query,
    delete_daily_case_list_tables_sql_query,
    select_case_data_sql_query,
    select_distinct_offense_statute_sql_query,
    select_distinct_def_last_def_first_case_number_sql_query,
    create_charges_table_sql_query,
)
from package.models.case_information import CriminalCaseInformation
from settings import DB_PATH, CHARGES_TABLE, EXCEL_DAILY_CASE_LISTS


class CaseSQLRetriever(ABC):
    @abstractmethod
    def query_case_data(self):
        raise NotImplementedError

    @abstractmethod
    def load_data_into_case(self):
        raise NotImplementedError

    @abstractmethod
    def load_case(self):
        raise NotImplementedError


class CriminalCaseSQLRetriever(CaseSQLRetriever):
    """Class for getting case data from a SQL database and packaging it into an object to be
    loaded into the application.

    :case_number: The selected case number from the case selected in the daily case lists box
    on the main window of the application.

    :case_table: The string name of the daily case list box that the case is selected from. This
    is passed to allow for certain options in the view to be changed based on which case list
    is selected. TODO: This can be refactored somehow."""

    def __init__(self, case_number: str, case_table: str) -> None:
        self.case_number = case_number
        self.case_table = case_table
        self.abbreviation_list = ["DUS", "OVI", "BMV"]
        self.database = open_db_connection("con_daily_case_lists")
        self.case = CriminalCaseInformation()
        self.query_case_data()
        self.load_data_into_case()
        self.query.finish()

    def query_case_data(self) -> None:
        """Query database based on cms_case number to return the data to load for the dialog.
        Current - Query.value(0) is id, then 1 is case_number, 2 is last_name, 3 is first_name."""
        query_string = select_case_data_sql_query(self.case_table, self.case_number)
        self.query = QSqlQuery(self.database)
        self.query.prepare(query_string)
        self.query.bindValue(self.case_number, self.case_number)
        self.query.exec()

    def load_data_into_case(self) -> None:
        while self.query.next():
            if self.case.case_number is None:
                self.load_case_information()
            self.load_charge_information()

    def load_case_information(self) -> None:
        self.case.case_number = self.query.value(1)
        self.case.defendant.last_name = self.query.value(2).title()
        self.case.defendant.first_name = self.query.value(3).title()
        self.case.fra_in_file = self.query.value(7)
        self.case.defense_counsel = f"{self.query.value(10).title()} {self.query.value(9).title()}"
        self.case.defense_counsel_type = self.query.value(11)

    def load_charge_information(self) -> None:
        offense = self.clean_offense_name(self.query.value(4))
        statute = self.query.value(5)
        degree = self.query.value(6)
        moving_bool = self.query.value(8)
        charge = (offense, statute, degree, moving_bool)
        self.case.charges_list.append(charge)

    def clean_offense_name(self, offense: str) -> str:
        """Sets an offense name to title case, but leaves certain standard 3-letter
        abbreviations in all caps."""
        if offense[:3] in self.abbreviation_list:
            caps = offense[:3]
            remaining_offense = string.capwords(offense[3:]).rstrip()
            return f"{caps} {remaining_offense}"
        return string.capwords(offense).rstrip()

    def load_case(self) -> CriminalCaseInformation:
        return self.case


def open_db_connection(connection_name: str) -> QSqlDatabase:
    db_connection = QSqlDatabase.database(connection_name, open=True)
    check_if_db_open(db_connection, connection_name)
    return db_connection


def remove_db_connection(connection_name: str) -> None:
    QSqlDatabase.removeDatabase(connection_name)


def create_db_connection(database_name: str, connection_name: str) -> QSqlDatabase:
    if not os.path.exists(database_name):
        print("The database does not exist. Creating new database.")
    db_connection = create_db(database_name, connection_name)
    return db_connection


def create_db(database_name: str, connection_name: str) -> QSqlDatabase:
    db_connection = QSqlDatabase.addDatabase("QSQLITE", connection_name)
    db_connection.setDatabaseName(database_name)
    return db_connection


def check_if_db_open(db_connection: QSqlDatabase, connection_name: str) -> bool:
    if not db_connection.isOpen():
        print(f"Unable to connect to {connection_name} database")
        sys.exit(1)
    return True


def create_daily_case_list_sql_tables(con_daily_case_lists: QSqlDatabase) -> None:
    for item in EXCEL_DAILY_CASE_LISTS:
        table_name: str
        table_name = item[1]
        QSqlQuery(con_daily_case_lists).exec(create_daily_case_list_tables_sql_query(table_name))


def load_daily_case_list_data(con_daily_case_lists: QSqlDatabase) -> None:
    for item in EXCEL_DAILY_CASE_LISTS:
        excel_report: str
        table_name: str
        excel_report, table_name = item
        delete_existing_daily_case_list_sql_table(con_daily_case_lists, table_name)
        insert_daily_case_list_sql_data(con_daily_case_lists, excel_report, table_name)


def insert_daily_case_list_sql_data(
    con_daily_case_lists: QSqlDatabase, excel_report: str, table_name: str
) -> None:
    cases_from_table = return_cases_data_from_excel(f"{DB_PATH}{excel_report}")
    for case in cases_from_table:
        insert_data_query = QSqlQuery(con_daily_case_lists)
        insert_data_query.prepare(insert_daily_case_list_tables_sql_query(table_name, case))
        insert_data_query.exec()


def delete_existing_daily_case_list_sql_table(
    con_daily_case_lists: QSqlDatabase, table_name: str
) -> None:
    delete_daily_case_list_table = QSqlQuery(con_daily_case_lists)
    delete_daily_case_list_table.prepare(delete_daily_case_list_tables_sql_query(table_name))
    delete_daily_case_list_table.exec()


class CaseExcelRetriever:
    def __init__(self) -> None:
        self.case_number: str = "No Data"
        self.sub_case_number: str = "No Data"
        self.defendant_last_name: str = "No Data"
        self.defendant_first_name: str = "No Data"
        self.offense: str = "No Data"
        self.statute: str = "No Data"
        self.degree: str = "No Data"
        self.fra_in_file: str = "U"
        self.moving_bool: str = "No Data"
        self.def_atty_last_name: str = ""
        self.def_atty_first_name: str = ""
        self.def_atty_type: str = "No Data"


def return_cases_data_from_excel(excel_file: str) -> list:
    col_case_number = 1
    col_sub_case_number = 2
    col_defendant_last_name = 3
    col_defendant_first_name = 4
    col_offense = 5
    col_statute = 6
    col_degree = 7
    col_fra_in_file = 8
    col_moving_bool = 9
    col_def_atty_last_name = 10
    col_def_atty_first_name = 11
    col_def_atty_type = 12
    data: list = []

    workbook = load_workbook(excel_file)
    ws = workbook.active
    row_count = ws.max_row + 1

    for row in range(2, row_count):
        case = CaseExcelRetriever()
        case.case_number = get_cell_value(ws, row, col_case_number)
        case.defendant_last_name = get_cell_value(ws, row, col_defendant_last_name)
        case.defendant_first_name = get_cell_value(ws, row, col_defendant_first_name)
        case.offense = get_cell_value(ws, row, col_offense)
        case.statute = get_cell_value(ws, row, col_statute)
        case.degree = get_cell_value(ws, row, col_degree)
        case.fra_in_file = get_cell_value(ws, row, col_fra_in_file)
        case.moving_bool = get_cell_value(ws, row, col_moving_bool)
        case.def_atty_last_name = get_cell_value(ws, row, col_def_atty_last_name)
        case.def_atty_first_name = get_cell_value(ws, row, col_def_atty_first_name)
        case.def_atty_type = get_cell_value(ws, row, col_def_atty_type)
        data.append(case)

    return data


def get_cell_value(ws: object, row: int, col: int) -> str:
    if ws.cell(row=row, column=col).value is None:
        if col == 8:  # fra_in_file
            return "U"
        if col == 10:
            return ""
        if col == 11:  # def_atty_last_name and def_atty_first_name
            return ""
        return "No Data"
    return ws.cell(row=row, column=col).value


def query_offense_statute_data(query_value: str) -> list:
    query_int = get_offense_statute_query_int(query_value)
    conn = open_db_connection("con_charges")
    query_string = select_distinct_offense_statute_sql_query()
    query = QSqlQuery(conn)
    query.prepare(query_string)
    query.exec()
    item_list = []
    while query.next():
        item_list.append(query.value(query_int))
    item_list.sort()
    conn.close()
    return item_list


def get_offense_statute_query_int(query_value: str) -> int:
    if query_value == "offense":
        return 0
    if query_value == "statute":
        return 1


def query_daily_case_list_data(table: str) -> list:
    conn = open_db_connection("con_daily_case_lists")
    query_string = select_distinct_def_last_def_first_case_number_sql_query(table)
    query = QSqlQuery(conn)
    query.prepare(query_string)
    query.exec()
    item_list = []
    while query.next():
        last_name = query.value(0).title()
        case_number = query.value(2)
        name = f"{last_name} - {case_number}"
        item_list.append(name)
    item_list.sort()
    conn.close()
    item_list.insert(0, None)
    return item_list




def update_charges_db(con_charges):
    create_charges_sql_table(con_charges)

    insertDataQuery = QSqlQuery(con_charges)
    insertDataQuery.prepare(
        """
        INSERT INTO charges (
            offense,
            statute,
            degree,
            type
        )
        VALUES (?, ?, ?, ?)
        """
    )
    # TO POPULATE A COMBO BOX
    # http://www.voidynullness.net/blog/2013/02/05/qt-populate-combo-box-from-database-table/
    # https://python-forum.io/thread-11659.html
    # Create two tables one for alpha sort and one for num sort - defintely a better way to do this
    data_from_table = return_charges_data_from_excel(CHARGES_TABLE)
    # print(data_from_table)
    # Use .addBindValue() to insert data
    for offense, statute, degree, type in data_from_table:
        insertDataQuery.addBindValue(offense)
        insertDataQuery.addBindValue(statute)
        insertDataQuery.addBindValue(degree)
        insertDataQuery.addBindValue(type)
        insertDataQuery.exec()
    con_charges.close()


def create_charges_sql_table(con_charges: str) -> None:
    QSqlQuery(con_charges).exec(create_charges_table_sql_query())


def return_charges_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    max_row = worksheet.max_row
    max_row = max_row + 1
    for row in range(2, max_row):
        offense = worksheet.cell(row=row, column=1)
        statute = worksheet.cell(row=row, column=2)
        degree = worksheet.cell(row=row, column=3)
        type = worksheet.cell(row=row, column=4)
        charge = (offense.value, statute.value, degree.value, type.value)
        data.append(charge)
    return data


def sql_query_offense_type(key):
    charges_database = open_db_connection("con_charges")
    query = QSqlQuery(charges_database)
    query.prepare("SELECT * FROM charges WHERE statute LIKE '%' || :key || '%'")
    query.bindValue(":key", key)
    query.exec()
    while query.next():
        statute = query.value(2)
        offense_type = query.value(4)
        if statute == key:
            query.finish()
            charges_database.close()
            return offense_type


def extract_data(case_data):
    wb_name = "Case_Data.xlsx"
    wb_name = DB_PATH + wb_name
    wb = load_workbook(wb_name)
    page = wb.active
    case_number = case_data.get("case_number")
    judicial_officer = case_data.get("judicial_officer").last_name
    charges_list = case_data.get("charges_list")
    max_row = page.max_row
    max_row = max_row + 1
    for index, charge in enumerate(charges_list):
        page.cell(row=max_row + index, column=1, value=case_number)
        page.cell(row=max_row + index, column=2, value=judicial_officer)
        page.cell(row=max_row + index, column=3, value=charge.get("offense"))
        page.cell(row=max_row + index, column=4, value=charge.get("statute"))
        page.cell(row=max_row + index, column=5, value=charge.get("degree"))
        page.cell(row=max_row + index, column=6, value=charge.get("plea"))
        page.cell(row=max_row + index, column=7, value=charge.get("finding"))
        page.cell(row=max_row + index, column=8, value=charge.get("fines_amount"))
        page.cell(row=max_row + index, column=9, value=charge.get("fines_suspended"))
        page.cell(row=max_row + index, column=10, value=charge.get("jail_days"))
        page.cell(row=max_row + index, column=11, value=charge.get("jail_days_suspended"))
    try:
        wb.save(filename=wb_name)
    except PermissionError:
        pass


def main():
    create_db_connection(f"{DB_PATH}daily_case_lists.sqlite", "con_daily_case_lists")
    con_daily_case_lists = open_db_connection("con_daily_case_lists")
    create_daily_case_list_sql_tables(con_daily_case_lists)
    load_daily_case_list_data(con_daily_case_lists)

    create_db_connection(f"{DB_PATH}charges.sqlite", "con_charges")
    con_charges = open_db_connection("con_charges")
    update_charges_db(con_charges)

    return None


if __name__ == "__main__":
    print("Daily Case Lists and Charges Tables created directly from script")
else:
    main()
    print("Imported Daily Case Lists and Charges Tables")
