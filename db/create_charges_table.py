"""This module is used to create the charges table."""
import sys
import os

from loguru import logger
from openpyxl import load_workbook
from PyQt5.QtSql import QSqlDatabase, QSqlQuery

from db.databases import open_db_connection, create_db_connection
from settings import CHARGES_TABLE, CHARGES_DATABASE, DB_PATH


@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    max_row = worksheet.max_row
    max_row = max_row + 1
    for row in range(2, max_row):
        offense = worksheet.cell(row=row, column=1)
        statute = worksheet.cell(row=row, column=2)
        degree = worksheet.cell(row=row, column=3)
        type = worksheet.cell(row=row, column=4)
        charge = (offense.value, statute.value, degree.value, type.value)
        data.append(charge)
    return data


def update_charges_db(con_charges):
    createTableQuery = QSqlQuery(con_charges)
    createTableQuery.exec(
        """
        CREATE TABLE charges (
            id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
            offense VARCHAR(60) UNIQUE NOT NULL,
            statute VARCHAR(50) NOT NULL,
            degree VARCHAR(50) NOT NULL,
            type VARCHAR(50) NOT NULL
        )
        """
    )
    insertDataQuery = QSqlQuery(con_charges)
    insertDataQuery.prepare(
        """
        INSERT INTO charges (
            offense,
            statute,
            degree,
            type
        )
        VALUES (?, ?, ?, ?)
        """
    )
    # TO POPULATE A COMBO BOX http://www.voidynullness.net/blog/2013/02/05/qt-populate-combo-box-from-database-table/
    # https://python-forum.io/thread-11659.html
    # Create two tables one for alpha sort and one for num sort - defintely a better way to do this
    data_from_table = return_data_from_excel(CHARGES_TABLE)
    # print(data_from_table)
    # Use .addBindValue() to insert data
    for offense, statute, degree, type in data_from_table:
        insertDataQuery.addBindValue(offense)
        insertDataQuery.addBindValue(statute)
        insertDataQuery.addBindValue(degree)
        insertDataQuery.addBindValue(type)
        insertDataQuery.exec()
    con_charges.close()


def main():
    create_db_connection(f"{DB_PATH}charges.sqlite", "con_charges")
    con_charges = open_db_connection("con_charges")
    update_charges_db(con_charges)
    return None


if __name__ == "__main__":
    print("Charges Table created directly from script")
else:
    main()
    print("Imported Charges Table")