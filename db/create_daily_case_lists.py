"""This module is used to create the slated table."""
import sys
import os
import pathlib
from loguru import logger

from openpyxl import load_workbook

from PyQt5.QtSql import QSqlDatabase, QSqlQuery

from settings import DB_PATH



@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    row_count = worksheet.max_row + 1
    for row in range(2, row_count):
        case_number = worksheet.cell(row=row, column=1)
        defendant_last_name = worksheet.cell(row=row, column=3)
        defendant_first_name = worksheet.cell(row=row, column=4)
        offense = worksheet.cell(row=row, column=5)
        if worksheet.cell(row=row, column=6).value is None:
            worksheet.cell(row=row, column=6).value = "No Data"
            statute = worksheet.cell(row=row, column=6)
        else:
            statute = worksheet.cell(row=row, column=6)
        if worksheet.cell(row=row, column=7).value is None:
            worksheet.cell(row=row, column=7).value = "No Data"
            degree = worksheet.cell(row=row, column=7)
        else:
            degree = worksheet.cell(row=row, column=7)
        if worksheet.cell(row=row, column=8).value is None:
            worksheet.cell(row=row, column=8).value = "U"
            fra_in_file = worksheet.cell(row=row, column=8)
        else:
            fra_in_file = worksheet.cell(row=row, column=8)
        case = (case_number.value,
                defendant_last_name.value,
                defendant_first_name.value,
                offense.value,
                statute.value,
                degree.value,
                fra_in_file.value,
                )
        data.append(case)
    return data


def main():
    """database_list contains tuples for the items to create each database for the daily case
    lists. Tuple format is (Excel_file_to_load, database_path_name, database_connection_name).

    The try execept block in the for items in database_list loop is to account for multiple users accessing the
    application at the same time. It creates a backup database with the information from the excel files that have
    the case information. This is done so that charges aren't duplicated because right now there is a PermissionError
    when the application is open if a second instance of the application is opened because it can't access the sqlite
    db files. TODO: Look into just creating a second connection without having to create a backup copy of the
    database."""

    # database_list = [
    #     ("Slated.xlsx", "slated.sqlite", "slated_table"),
    #     ("Arraignments.xlsx", "arraignments.sqlite", "arraignments_table"),
    #     ("Final_Pretrials.xlsx", "final_pretrials.sqlite", "final_pretrials_table"),
    # ]

    database_table = "arraignments"
    database_name = f"{DB_PATH}arraignments.sqlite"
    excel_report = "Arraignments.xlsx"

    if os.path.exists(database_name):
        con1 = QSqlDatabase.addDatabase("QSQLITE", "con1")
        con1.setDatabaseName(database_name)
    else:
        print("The file does not exist")
        con1 = QSqlDatabase.addDatabase("QSQLITE", "con1")
        con1.setDatabaseName(database_name)
        if not con1.open():
            print("Unable to connect to database")
            sys.exit(1)
        else:
            create_table_query = QSqlQuery(con1)
            create_table_query.exec(
                """
                CREATE TABLE arraignments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
                    case_number VARCHAR(20) NOT NULL,
                    defendant_last_name VARCHAR(50) NOT NULL,
                    defendant_first_name VARCHAR(50) NOT NULL,
                    offense VARCHAR(80) NOT NULL,
                    statute VARCHAR(50) NOT NULL,
                    degree VARCHAR(10) NOT NULL,
                    fra_in_file VARCHAR(5) NOT NULL
                )
                """
            )

    if not con1.open():
        print("Unable to connect to database")
        sys.exit(1)

    delete_old_data_query = QSqlQuery(con1)
    delete_old_data_query.prepare(
        """
        DELETE FROM arraignments;
        """
    )
    #delete_old_data_query.bindValue(database_table, database_table)
    delete_old_data_query.exec()

    insert_data_query = QSqlQuery(con1)
    # Do not add comma to last value inserted
    insert_data_query.prepare(
        """
        INSERT INTO arraignments(
            case_number,
            defendant_last_name,
            defendant_first_name,
            offense,
            statute,
            degree,
            fra_in_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """
    )
    # insert_data_query.bindValue(database_table, database_table)

    data_from_table = return_data_from_excel(f"{DB_PATH}{excel_report}")
    # Use .addBindValue() to insert data
    for case_number, defendant_last_name, defendant_first_name, offense, \
            statute, degree, fra_in_file in data_from_table:
        insert_data_query.addBindValue(case_number)
        insert_data_query.addBindValue(defendant_last_name)
        insert_data_query.addBindValue(defendant_first_name)
        insert_data_query.addBindValue(offense)
        insert_data_query.addBindValue(statute)
        insert_data_query.addBindValue(degree)
        insert_data_query.addBindValue(fra_in_file)
        insert_data_query.exec()
    con1.close()
    con1.removeDatabase("QSQLITE")
    return None


if __name__ == "__main__":
    print("Daily Case Lists created directly from script")
else:
    main()
    print("Imported Daily Case List Tables")
