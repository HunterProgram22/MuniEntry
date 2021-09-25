import sys
import pathlib
from loguru import logger

from openpyxl import load_workbook

from PyQt5.QtSql import QSqlDatabase, QSqlQuery


PATH = str(pathlib.Path().absolute())

EXCEL_FILE = PATH + "\Case_Types.xlsx"

print(EXCEL_FILE)


@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    for row in range(2, 30):
        offense = worksheet.cell(row=row, column=1)
        statute = worksheet.cell(row=row, column=2)
        degree = worksheet.cell(row=row, column=3)
        charge = (offense.value, statute.value, degree.value)
        data.append(charge)
    return data


con = QSqlDatabase.addDatabase("QSQLITE")
con.setDatabaseName(PATH + "\\charges.sqlite")

if not con.open():
    print("Unable to connect to database")
    sys.exit(1)

# Create a query and execute it right away using .exec()
createTableQuery = QSqlQuery()
createTableQuery.exec(
    """
    CREATE TABLE charges (
        id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
        offense VARCHAR(60) NOT NULL,
        statute VARCHAR(50) NOT NULL,
        degree VARCHAR(50) NOT NULL
    )
    """
)

print(con.tables())

insertDataQuery = QSqlQuery()
insertDataQuery.prepare(
    """
    INSERT INTO charges (
        offense,
        statute,
        degree
    )
    VALUES (?, ?, ?)
    """
)


# TO POPULATE A COMBO BOX http://www.voidynullness.net/blog/2013/02/05/qt-populate-combo-box-from-database-table/
# https://python-forum.io/thread-11659.html
# Sample data
data_from_table = return_data_from_excel(EXCEL_FILE)
print(data_from_table)

# Use .addBindValue() to insert data
for offense, statute, degree in data_from_table:
    insertDataQuery.addBindValue(offense)
    insertDataQuery.addBindValue(statute)
    insertDataQuery.addBindValue(degree)
    insertDataQuery.exec()
