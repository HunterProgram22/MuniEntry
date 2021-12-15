"""This module is used to create the arraignments table."""
import sys
import os
import pathlib
from loguru import logger

from openpyxl import load_workbook

from PyQt5.QtSql import QSqlDatabase, QSqlQuery


PATH = str(pathlib.Path().absolute())

"""This path is from the MuniEntry level of the program because it uses Path at the level of the module
into which it is imported."""
EXCEL_FILE = PATH + "\\resources\\db\\Arraignments.xlsx"


@logger.catch
def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    row_count = worksheet.max_row + 1
    for row in range(2,row_count):
        case_number = worksheet.cell(row=row, column=1)
        defendant_last_name = worksheet.cell(row=row, column=3)
        defendant_first_name = worksheet.cell(row=row, column=4)
        offense = worksheet.cell(row=row, column=5)
        statute = worksheet.cell(row=row, column=6)
        degree = worksheet.cell(row=row, column=7)
        if worksheet.cell(row=row, column=8).value == None:
            worksheet.cell(row=row, column=8).value = "U"
            fra_in_file = worksheet.cell(row=row, column=8)
        else:
            fra_in_file = worksheet.cell(row=row, column=8)
        case = (case_number.value,
                defendant_last_name.value,
                defendant_first_name.value,
                offense.value,
                statute.value,
                degree.value,
                fra_in_file.value,
                )
        data.append(case)
    return data

def main():
    if os.path.exists(PATH + "\\resources\db\\arraignments.sqlite"):
        os.remove(PATH + "\\resources\db\\arraignments.sqlite")
    else:
        print("The file does not exist")
    con = QSqlDatabase.addDatabase("QSQLITE")
    con.setDatabaseName(PATH + "\\resources\db\\arraignments.sqlite")

    if not con.open():
        print("Unable to connect to database")
        sys.exit(1)

    # Create a query and execute it right away using .exec()
    createTableQuery = QSqlQuery()
    createTableQuery.exec(
        """
        CREATE TABLE cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL,
            case_number VARCHAR(20) NOT NULL,
            defendant_last_name VARCHAR(50) NOT NULL,
            defendant_first_name VARCHAR(50) NOT NULL,
            offense VARCHAR(80) NOT NULL,
            statute VARCHAR(50) NOT NULL,
            degree VARCHAR(10) NOT NULL,
            fra_in_file VARCHAR(5) NOT NULL
        )
        """
    )
    insertDataQuery = QSqlQuery()
    # Do not add comma to last value inserted
    insertDataQuery.prepare(
        """
        INSERT INTO cases (
            case_number,
            defendant_last_name,
            defendant_first_name,
            offense,
            statute,
            degree,
            fra_in_file
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """
    )
    # TO POPULATE A COMBO BOX http://www.voidynullness.net/blog/2013/02/05/qt-populate-combo-box-from-database-table/
    # https://python-forum.io/thread-11659.html
    # Create two tables one for alpha sort and one for num sort - defintely a better way to do this
    data_from_table = return_data_from_excel(EXCEL_FILE)
    # print(data_from_table)

    # Use .addBindValue() to insert data
    for case_number, defendant_last_name, defendant_first_name, offense, statute, degree, fra_in_file in data_from_table:
        insertDataQuery.addBindValue(case_number)
        insertDataQuery.addBindValue(defendant_last_name)
        insertDataQuery.addBindValue(defendant_first_name)
        insertDataQuery.addBindValue(offense)
        insertDataQuery.addBindValue(statute)
        insertDataQuery.addBindValue(degree)
        insertDataQuery.addBindValue(fra_in_file)
        insertDataQuery.exec()
    return None

if __name__ == "__main__":
    print("Create arraignments table ran directly")
else:
    main()
    print("Create arraignments table ran when imported")
